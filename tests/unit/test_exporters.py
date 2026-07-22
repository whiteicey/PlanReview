from __future__ import annotations

import json
import zipfile

from docx import Document
from openpyxl import load_workbook

from app.domain.enums import Origin, ReviewStatus, RuleStatus, Severity
from app.domain.schemas import Finding, RuleResult
from app.reports.exporters import export_anonymous_package, export_excel, export_word, safe_excel_cell
from app.review.pipeline import ReviewRun


def _run() -> ReviewRun:
    return ReviewRun(
        "CASE-1",
        findings=[
            Finding(
                finding_id="F1",
                origin=Origin.RULE,
                category="capacity",
                severity=Severity.HIGH,
                title="Capacity mismatch",
                description="The proposed capacity needs confirmation.",
                suggestion="Confirm the supporting calculation.",
                evidence_span_ids=["span-1"],
                needs_human_review=True,
                review_status=ReviewStatus.CONFIRMED,
                human_note="专家已确认",
            )
        ],
        rule_results=[
            RuleResult(
                rule_id="CAP-001",
                status=RuleStatus.FAIL,
                severity=Severity.HIGH,
                category="capacity",
                details={"rule_version": "2026.07"},
            )
        ],
        evidence_text_hashes={"span-1": "a" * 64},
    )


def test_excel_and_word_exports_include_disclaimer_evidence_and_review_state(tmp_path):
    run = _run()
    spreadsheet = export_excel(run, tmp_path / "reports" / "findings.xlsx")
    document = export_word(run, tmp_path / "reports" / "findings.docx")

    assert spreadsheet.exists() and document.exists()
    sheet = load_workbook(spreadsheet).active
    assert "不是正式审查结论" in sheet["A1"].value
    headers = [cell.value for cell in sheet[2]]
    exported = dict(zip(headers, [cell.value for cell in sheet[3]], strict=True))
    assert exported["evidence_span_ids"] == "span-1"
    assert exported["review_status"] == "confirmed"
    assert exported["human_note"] == "专家已确认"

    text = "\n".join(paragraph.text for paragraph in Document(document).paragraphs)
    assert "不是正式审查结论" in text
    assert "span-1" in text
    assert "confirmed" in text
    assert "专家已确认" in text


def test_anonymous_package_contains_only_honest_anonymized_export_data(tmp_path):
    target = export_anonymous_package(_run(), tmp_path / "anonymous.zip")

    with zipfile.ZipFile(target) as archive:
        assert archive.namelist() == ["anonymous-findings.json"]
        payload = json.loads(archive.read("anonymous-findings.json"))

    assert "不是正式审查结论" in payload["disclaimer"]
    assert "case_id" not in payload
    assert payload["rule_versions"] == [{"rule_id": "rule-0001", "version": "version-0001"}]
    assert payload["evidence_text_hashes"] == {"evidence-0001": "a" * 64}
    assert payload["metrics"] == {
        "finding_count": 1,
        "review_state_counts": {"reviewstatus-0002": 1},
        "accuracy": "not_measured",
        "recall": "not_measured",
        "time_saved": "not_measured",
        "cost": "not_measured",
    }
    assert payload["findings"][0]["review_status"] == "reviewstatus-0002"
    assert payload["findings"][0]["category"] == "category-0005"
    assert payload["findings"][0]["severity"] == "severity-0001"
    assert "title" not in payload["findings"][0]
    assert "description" not in payload["findings"][0]
    assert "suggestion" not in payload["findings"][0]
    assert "human_note" not in payload["findings"][0]


def test_excel_export_never_creates_formulas_from_external_text(tmp_path):
    attacks = [
        "=1+1",
        '=HYPERLINK("https://example.com","click")',
        "+CMD",
        "-2+3",
        "@SUM(A1:A2)",
    ]
    run = _run()
    run.findings[0] = run.findings[0].model_copy(
        update={
            "title": attacks[0],
            "description": attacks[1],
            "suggestion": attacks[2],
            "human_note": attacks[3],
            "category": attacks[4],
        }
    )
    run.rule_results[0] = run.rule_results[0].model_copy(
        update={"message": attacks[0], "category": attacks[1], "evidence_span_ids": ["span-1"]}
    )
    run.evidence_locations = {"span-1": attacks[2]}
    target = export_excel(
        run,
        tmp_path / "safe.xlsx",
        evidence_texts={"span-1": attacks[3]},
        evidence_file_names={"span-1": attacks[4]},
    )

    workbook = load_workbook(target, data_only=False)
    assert all(
        cell.data_type != "f"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
    )
    assert all(safe_excel_cell(value).startswith("'") for value in attacks)


def test_safe_excel_cell_preserves_typed_scalars_and_removes_illegal_controls():
    from datetime import date

    values = [1, 2.5, True, date(2026, 1, 2), None]
    assert [safe_excel_cell(value) for value in values] == values
    assert safe_excel_cell("safe\x00text") == "safetext"
