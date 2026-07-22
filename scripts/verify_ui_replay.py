"""Verify a V1.2 UI replay against its source database without business writes."""

from __future__ import annotations

import argparse
import hashlib
import json
import sqlite3
import tempfile
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


def _json(value):
    return json.loads(value) if isinstance(value, str) else value


def snapshot(database: Path, run_id: str) -> dict:
    connection = sqlite3.connect(database)
    connection.row_factory = sqlite3.Row
    run = connection.execute("SELECT * FROM review_runs WHERE run_id = ?", (run_id,)).fetchone()
    if run is None:
        raise RuntimeError(f"run not found: {run_id}")
    findings = connection.execute(
        "SELECT finding_id, origin, severity, title, evidence_span_ids, review_status FROM findings WHERE review_run_id = ? ORDER BY position",
        (run["id"],),
    ).fetchall()
    rules = connection.execute("SELECT rule_id FROM rule_results WHERE review_run_id = ? ORDER BY position", (run["id"],)).fetchall()
    batch_metrics = _json(run["batch_metrics"]) or []
    packet_ledger = _json(run["packet_lifecycle_ledger"]) or {}
    candidate_ledger = _json(run["ai_candidate_lifecycle_ledger"]) or {}
    stage_records = _json(run["stage_records"]) or []
    normalized = [
        {
            "id": row["finding_id"],
            "origin": row["origin"],
            "severity": row["severity"],
            "title": row["title"],
            "source": sorted(_json(row["evidence_span_ids"]) or []),
            "review_status": row["review_status"],
        }
        for row in findings
    ]
    finding_hash = hashlib.sha256(json.dumps(normalized, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode()).hexdigest()
    result = {
        "run_count": connection.execute("SELECT COUNT(*) FROM review_runs").fetchone()[0],
        "final_status": run["final_status"],
        "llm_status": run["llm_status"],
        "finding_count": len(findings),
        "finding_ids": [row["finding_id"] for row in findings],
        "finding_core_sha256": finding_hash,
        "rule_result_count": len(rules),
        "distinct_rule_ids": sorted({row["rule_id"] for row in rules}),
        "batch_metrics": [
            {"batch_id": item.get("batch_id"), "batch_index": item.get("batch_index"), "stop_reason": item.get("stop_reason")}
            for item in batch_metrics
        ],
        "packet_ledger": {key: packet_ledger.get(key) for key in ("ledger_entry_count", "ledger_size_bytes", "ledger_truncated")},
        "candidate_ledger": {key: candidate_ledger.get(key) for key in ("ledger_entry_count", "ledger_size_bytes", "ledger_truncated")},
        "stage_records": [{"stage": item.get("stage"), "status": item.get("status")} for item in stage_records],
    }
    connection.close()
    return result


def get_json(base_url: str, path: str):
    with urllib.request.urlopen(f"{base_url}{path}", timeout=20) as response:
        return json.loads(response.read())


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-db", type=Path, required=True)
    parser.add_argument("--replay-db", type=Path, required=True)
    parser.add_argument("--base-url", required=True)
    parser.add_argument("--case-id", required=True)
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    before = snapshot(args.source_db, args.run_id)
    after = snapshot(args.replay_db, args.run_id)
    summary = get_json(args.base_url, f"/api/cases/{args.case_id}/runs/{args.run_id}")
    findings = get_json(args.base_url, f"/api/cases/{args.case_id}/runs/{args.run_id}/findings")
    diagnostics = get_json(args.base_url, f"/api/cases/{args.case_id}/runs/{args.run_id}/diagnostics")
    with tempfile.TemporaryDirectory(prefix="planreview-ui-export-") as temp:
        export_path = Path(temp) / "review.xlsx"
        urllib.request.urlretrieve(f"{args.base_url}/api/cases/{args.case_id}/exports/xlsx", export_path)
        workbook = load_workbook(export_path, read_only=True, data_only=True)
        # Findings has a disclaimer row and a column-header row; Rules has one header row.
        xlsx_findings = max(0, workbook["Findings"].max_row - 2)
        xlsx_rules = max(0, workbook["Rules"].max_row - 1)
        workbook.close()
    checks = {
        "database_snapshot_unchanged": before == after,
        "api_finding_count": len(findings) == 25 == summary.get("finding_count"),
        "diagnostics_counts": diagnostics.get("integrity") == {
            "packet_ledger_entries": 504,
            "candidate_ledger_entries": 119,
            "packet_ledger_truncated": False,
            "candidate_ledger_truncated": False,
            "batch_count": 5,
            "finding_count": 25,
            "rule_result_count": 29,
            "distinct_rule_id_count": 18,
        },
        "xlsx_counts": xlsx_findings == 25 and xlsx_rules == 29,
        "stages_completed_in_database": len(after["stage_records"]) == 8 and all(item["status"] == "completed" for item in after["stage_records"]),
    }
    report = {
        "run_id": "<redacted-run-id>",
        "case_id": "<redacted-case-id>",
        "source": before,
        "replay_after": after,
        "api": {"finding_count": len(findings), "final_status": summary.get("final_status"), "llm_status": summary.get("llm_status"), "diagnostics_integrity": diagnostics.get("integrity")},
        "xlsx": {"finding_rows": xlsx_findings, "rule_rows": xlsx_rules},
        "checks": checks,
        "passed": all(checks.values()),
    }
    report["source"]["finding_ids"] = [f"<finding-{index:02d}>" for index, _ in enumerate(report["source"]["finding_ids"], 1)]
    report["replay_after"]["finding_ids"] = list(report["source"]["finding_ids"])
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")
    print(json.dumps({"passed": report["passed"], "checks": checks}, ensure_ascii=False))
    return 0 if report["passed"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
